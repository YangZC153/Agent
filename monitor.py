#!/usr/bin/env python3
"""
arxiv 生成式推荐论文自动监控主脚本
每天定时执行：搜索 -> 查重 -> 下载 PDF -> 输出结构化 JSON
中文总结和作者单位提取由 hermes cronjob agent 调用 LLM 完成
"""

import os
import sys
import json
import re
import time
import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ==================== 配置区 ====================
BASE_DIR = Path(__file__).resolve().parent
PAPERS_DIR = BASE_DIR / "papers"
EXCEL_FILE = BASE_DIR / "papers_record.xlsx"
VIEWER_JSON = BASE_DIR / "viewer" / "papers_data.json"
CRAWLED_IDS_FILE = BASE_DIR / "crawled_ids.txt"
PENDING_LLM_IDS_FILE = BASE_DIR / "pending_llm_ids.txt"
TOPICS_FILE = BASE_DIR / "search_topics.json"
OUTPUT_JSON = BASE_DIR / "new_papers.json"   # 输出给 hermes agent 的中间文件

# arxiv API 配置
ARXIV_API = "https://export.arxiv.org/api/query"
DAILY_LLM_LIMIT = 10
CANDIDATES_PER_TOPIC = 50
SCREENING_CANDIDATES_PER_TOPIC = 25
SCREENING_CANDIDATE_LIMIT = 125
SCREENING_BATCH_SIZE = 20
RECENT_PAPER_DAYS = 30
REQUEST_INTERVAL = 3  # 秒
ARXIV_TIMEOUT = 30
ARXIV_API_ATTEMPTS = 2
ARXIV_USER_AGENT = (
    "hermes-arxiv-agent/1.0 "
    "(mailto:YangZC153@users.noreply.github.com)"
)
DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"
DEEPSEEK_SCREENING_MODEL = "deepseek-v4-flash"

EXCEL_COLUMNS = [
    "arxiv_id",
    "title",
    "authors",
    "affiliations",
    "published_date",
    "categories",
    "abstract",
    "summary_cn",
    "topic_name",
    "code_open_source",
    "code_url",
    "pdf_filename",
    "crawled_date",
    "notes",
]

DEFAULT_SEARCH_TOPICS = [
    {
        "id": 1,
        "name": "Multimodal generative recommendation",
        "target_quota": 6,
        "query": (
            "(all:multimodal AND all:generative AND "
            "(all:recommendation OR all:recommender))"
        ),
        "ranking_terms": [
            "multimodal generative recommendation",
            "multimodal recommendation",
            "multimodal recommender",
            "generative recommendation",
            "generative recommender",
        ],
        "required_term_groups": [
            ["multimodal"],
            ["generative"],
            ["recommendation", "recommender"],
        ],
    },
    {
        "id": 2,
        "name": "Multimodal recommendation",
        "target_quota": 1,
        "query": (
            "(all:multimodal AND "
            "(all:recommendation OR all:recommender))"
        ),
        "ranking_terms": [
            "multimodal recommendation",
            "multimodal recommender",
            "multimodal sequential recommendation",
            "multimodal representation",
        ],
        "required_term_groups": [
            ["multimodal"],
            ["recommendation", "recommender"],
        ],
    },
    {
        "id": 3,
        "name": "Collaborative filtering + generative recommendation",
        "target_quota": 1,
        "query": (
            "(all:\"collaborative filtering\" AND "
            "(all:generative OR all:\"large language model\" OR all:LLM))"
        ),
        "ranking_terms": [
            "collaborative filtering",
            "generative recommendation",
            "generative recommender",
            "large language model",
        ],
        "required_term_groups": [
            ["collaborative filtering"],
            [
                "generative recommendation",
                "generative recommender",
                "semantic id",
                "large language model",
                "llm",
                "diffusion",
                "autoregressive",
            ],
        ],
    },
    {
        "id": 4,
        "name": "LLM + recommendation",
        "target_quota": 1,
        "query": (
            "((all:LLM OR all:\"large language model\") AND "
            "(all:\"recommender system\" OR all:\"recommendation system\" "
            "OR all:\"recommendation model\"))"
        ),
        "ranking_terms": [
            "large language model",
            "llm",
            "recommendation",
            "recommender",
        ],
        "required_term_groups": [
            ["large language model", "llm"],
            [
                "recommender system",
                "recommendation system",
                "recommendation model",
                "recommendation task",
            ],
        ],
    },
    {
        "id": 5,
        "name": "Agentic recommender systems",
        "target_quota": 1,
        "query": (
            "((all:agentic OR all:\"LLM agent\" OR all:\"multi-agent\" "
            "OR all:\"recommendation agent\" OR all:\"recommender agent\") AND "
            "(all:\"recommender system\" OR all:\"recommendation system\" "
            "OR all:\"recommendation agent\" OR all:\"recommender agent\"))"
        ),
        "ranking_terms": [
            "agentic recommender",
            "agentic recommendation",
            "recommender agent",
            "recommendation agent",
            "multi-agent",
            "recommender system",
            "recommendation system",
        ],
        "required_term_groups": [
            ["agentic", "llm agent", "multi-agent", "recommendation agent", "recommender agent"],
            ["recommender system", "recommendation system", "recommendation agent", "recommender agent"],
        ],
    },
]

# ==================== 工具函数 ====================

def load_crawled_ids() -> set:
    if not CRAWLED_IDS_FILE.exists():
        return set()
    with open(CRAWLED_IDS_FILE, "r", encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())


def load_excel_ids() -> set:
    """从 Excel 的 Papers 表读取已有 arxiv_id。"""
    if not EXCEL_FILE.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        if "Papers" not in wb.sheetnames:
            return set()
        ws = wb["Papers"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return set()
        headers = [str(h) if h is not None else "" for h in header_row]
        if "arxiv_id" not in headers:
            return set()
        arxiv_col = headers.index("arxiv_id")
        ids = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[arxiv_col] if arxiv_col < len(row) else None
            if val:
                ids.add(str(val).strip())
        return ids
    except Exception as e:
        print(f"[WARN] Failed to load Excel IDs: {e}")
        return set()


def load_pending_llm_ids() -> set:
    if not PENDING_LLM_IDS_FILE.exists():
        return set()
    with open(PENDING_LLM_IDS_FILE, "r", encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())


def save_pending_llm_ids(ids: set[str] | list[str]):
    cleaned = sorted({str(x).strip() for x in ids if str(x).strip()})
    if not cleaned:
        if PENDING_LLM_IDS_FILE.exists():
            PENDING_LLM_IDS_FILE.unlink()
        return
    with open(PENDING_LLM_IDS_FILE, "w", encoding="utf-8") as f:
        for arxiv_id in cleaned:
            f.write(arxiv_id + "\n")


def save_crawled_ids_batch(new_ids: list[str]):
    """批量追加新 ID"""
    with open(CRAWLED_IDS_FILE, "a", encoding="utf-8") as f:
        for arxiv_id in new_ids:
            f.write(arxiv_id + "\n")


def load_search_topics() -> list[dict]:
    if not TOPICS_FILE.exists():
        return DEFAULT_SEARCH_TOPICS

    try:
        with open(TOPICS_FILE, "r", encoding="utf-8") as f:
            topics = json.load(f)
        if not isinstance(topics, list) or not topics:
            raise ValueError("expected a non-empty JSON list")
        required = {"id", "name", "target_quota", "query", "ranking_terms"}
        for topic in topics:
            missing = required - set(topic)
            if missing:
                raise ValueError(f"topic missing fields: {sorted(missing)}")
        return sorted(topics, key=lambda item: int(item["id"]))
    except Exception as e:
        print(f"[WARN] Failed to load {TOPICS_FILE.name}: {e}; using defaults")
        return DEFAULT_SEARCH_TOPICS


def search_arxiv_papers(query: str, max_results: int = CANDIDATES_PER_TOPIC) -> list[dict]:
    params = {
        "search_query": query,
        "max_results": max_results,
        "sortBy": "submittedDate",
        "sortOrder": "descending",
    }

    print(f"[INFO] Searching arxiv: {query}")
    last_error = None
    for attempt in range(1, ARXIV_API_ATTEMPTS + 1):
        try:
            response = requests.get(
                ARXIV_API,
                params=params,
                headers={"User-Agent": ARXIV_USER_AGENT},
                timeout=ARXIV_TIMEOUT,
            )
            response.raise_for_status()
            break
        except Exception as exc:
            last_error = exc
            if attempt >= ARXIV_API_ATTEMPTS:
                raise
            status_code = getattr(getattr(exc, "response", None), "status_code", None)
            retry_after = getattr(getattr(exc, "response", None), "headers", {}).get(
                "Retry-After"
            )
            if retry_after and retry_after.isdigit():
                delay = int(retry_after)
            elif status_code == 429:
                delay = 20 * attempt
            elif status_code in {500, 502, 503, 504}:
                delay = 10 * attempt
            else:
                delay = REQUEST_INTERVAL * attempt
            print(
                f"[WARN] arXiv query failed on attempt "
                f"{attempt}/{ARXIV_API_ATTEMPTS}: {exc}; retrying in {delay}s"
            )
            time.sleep(delay)
    else:
        raise RuntimeError(f"arXiv query failed: {last_error}")

    ns = {"a": "http://www.w3.org/2005/Atom"}
    root = ET.fromstring(response.content)

    papers = []
    for entry in root.findall("a:entry", ns):
        try:
            arxiv_id_full = entry.find("a:id", ns).text.strip().split("/abs/")[-1]
            arxiv_id = arxiv_id_full.split("v")[0]   # 剥离版本号

            title = entry.find("a:title", ns).text.strip().replace("\n", " ")
            authors = ", ".join(
                a.find("a:name", ns).text for a in entry.findall("a:author", ns)
            )
            summary = entry.find("a:summary", ns).text.strip().replace("\n", " ")
            published = entry.find("a:published", ns).text[:10]
            cats = ", ".join(c.get("term") for c in entry.findall("a:category", ns))
            pdf_url = f"https://arxiv.org/pdf/{arxiv_id_full}"

            papers.append({
                "arxiv_id": arxiv_id,
                "title": title,
                "authors": authors,
                "summary": summary,
                "published_date": published,
                "categories": cats,
                "pdf_url": pdf_url,
                "pdf_filename": f"{arxiv_id}.pdf",
                "pdf_local_path": str(PAPERS_DIR / f"{arxiv_id}.pdf"),
                "affiliations": "",   # ← 由 hermes LLM 从 PDF 提取
                "summary_cn": "",      # ← 由 hermes LLM 生成中文总结
            })
        except Exception as e:
            print(f"[WARN] Parse error: {e}")
            continue

    return papers


def topic_relevance_score(paper: dict, topic: dict) -> int:
    """Local score is used for candidate pre-filtering and API-failure fallback."""
    title = paper.get("title", "").casefold()
    abstract = paper.get("summary", "").casefold()
    score = 0
    for term in topic.get("ranking_terms", []):
        needle = str(term).casefold().strip()
        if not needle:
            continue
        score += title.count(needle) * 8
        score += abstract.count(needle) * 3
    return score


def paper_matches_topic(paper: dict, topic: dict) -> bool:
    text = f"{paper.get('title', '')} {paper.get('summary', '')}".casefold()
    for group in topic.get("required_term_groups", []):
        if not any(str(term).casefold() in text for term in group):
            return False
    return True


def paper_is_recent(paper: dict, today: date | None = None) -> bool:
    today = today or date.today()
    try:
        published = date.fromisoformat(str(paper.get("published_date", "")))
    except ValueError:
        return False
    return published >= today - timedelta(days=RECENT_PAPER_DAYS)


def select_topic_papers(
    candidates_by_topic: dict[int, list[dict]],
    topics: list[dict],
    limit: int,
) -> list[dict]:
    """
    Select unique papers without an LLM call.

    First reserve one paper per topic, then meet each target quota (6/1/1/1/1
    by default), and finally fill gaps in topic priority order.
    """
    if limit <= 0:
        return []

    ranked: dict[int, list[dict]] = {}
    for topic in topics:
        topic_id = int(topic["id"])
        papers = candidates_by_topic.get(topic_id, [])
        ranked[topic_id] = sorted(
            papers,
            key=lambda paper: (
                1 if paper.get("code_open_source") == "是" else 0,
                int(paper.get("screening_score", 0)),
                topic_relevance_score(paper, topic),
                paper.get("published_date", ""),
                paper.get("arxiv_id", ""),
            ),
            reverse=True,
        )

    selected: list[dict] = []
    selected_ids: set[str] = set()
    selected_per_topic = {int(topic["id"]): 0 for topic in topics}

    def take(topic: dict, count: int) -> int:
        topic_id = int(topic["id"])
        taken = 0
        for paper in ranked.get(topic_id, []):
            if len(selected) >= limit or taken >= count:
                break
            arxiv_id = paper["arxiv_id"]
            if arxiv_id in selected_ids:
                continue
            selected.append({
                **paper,
                "topic_id": topic_id,
                "topic_name": topic["name"],
                "selection_score": topic_relevance_score(paper, topic),
            })
            selected_ids.add(arxiv_id)
            selected_per_topic[topic_id] += 1
            taken += 1
        return taken

    # Coverage pass: when enough slots exist, every topic gets at least one.
    for topic in topics:
        if len(selected) >= limit:
            break
        take(topic, 1)

    # Target pass: topic 1 grows to six; the other topics retain one each.
    for topic in topics:
        if len(selected) >= limit:
            break
        topic_id = int(topic["id"])
        remaining = max(0, int(topic["target_quota"]) - selected_per_topic[topic_id])
        take(topic, remaining)

    # Fill missing quota from the highest-priority topic with available papers.
    for topic in topics:
        if len(selected) >= limit:
            break
        take(topic, limit - len(selected))

    return selected


def load_deepseek_api_key() -> str:
    key = os.environ.get("DEEPSEEK_API_KEY", "").strip()
    if key:
        return key

    env_file = Path.home() / ".hermes" / ".env"
    if not env_file.exists():
        return ""
    for line in env_file.read_text(encoding="utf-8").splitlines():
        if not line.startswith("DEEPSEEK_API_KEY="):
            continue
        return line.split("=", 1)[1].strip().strip("\"'")
    return ""


def extract_json_object(text: str) -> dict:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start < 0 or end <= start:
            raise
        return json.loads(cleaned[start:end + 1])


def build_screening_pool(
    candidates_by_topic: dict[int, list[dict]],
    topics: list[dict],
) -> list[dict]:
    """Keep the prompt bounded while preserving candidates from every topic."""
    by_id: dict[str, dict] = {}
    for topic in topics:
        topic_id = int(topic["id"])
        ranked = sorted(
            candidates_by_topic.get(topic_id, []),
            key=lambda paper: (
                topic_relevance_score(paper, topic),
                paper.get("published_date", ""),
                paper.get("arxiv_id", ""),
            ),
            reverse=True,
        )[:SCREENING_CANDIDATES_PER_TOPIC]
        for paper in ranked:
            arxiv_id = paper["arxiv_id"]
            if arxiv_id not in by_id:
                by_id[arxiv_id] = {**paper, "candidate_topic_ids": []}
            by_id[arxiv_id]["candidate_topic_ids"].append(topic_id)

    pool = list(by_id.values())
    pool.sort(
        key=lambda paper: (
            paper.get("published_date", ""),
            paper.get("arxiv_id", ""),
        ),
        reverse=True,
    )
    return pool[:SCREENING_CANDIDATE_LIMIT]


def fallback_screen_candidates(candidates: list[dict], topics: list[dict]) -> list[dict]:
    """Keep the monitor usable when the low-cost screening API is unavailable."""
    screened = []
    topics_by_id = {int(topic["id"]): topic for topic in topics}
    for paper in candidates:
        candidate_topic_ids = paper.get("candidate_topic_ids", [])
        if not candidate_topic_ids:
            continue
        topic_id = max(
            candidate_topic_ids,
            key=lambda value: topic_relevance_score(paper, topics_by_id[int(value)]),
        )
        abstract = paper.get("summary", "")
        url_match = re.search(
            r"https?://(?:www\.)?(?:github\.com|gitlab\.com|codeberg\.org)/[^\s,;)\]]+",
            abstract,
            flags=re.IGNORECASE,
        )
        open_source = bool(
            url_match
            or re.search(
                r"\b(?:code|source code|implementation)\b.{0,80}"
                r"\b(?:open[- ]source|publicly available|available at|released)\b",
                abstract,
                flags=re.IGNORECASE,
            )
        )
        screened.append({
            **paper,
            "topic_id": int(topic_id),
            "topic_name": topics_by_id[int(topic_id)]["name"],
            "screening_score": topic_relevance_score(
                paper, topics_by_id[int(topic_id)]
            ),
            "code_open_source": "是" if open_source else "摘要未说明",
            "code_url": url_match.group(0).rstrip(".") if url_match else "",
        })
    return screened


def _screen_candidate_batch_with_deepseek(
    candidates: list[dict],
    topics: list[dict],
) -> list[dict]:
    if not candidates:
        return []

    api_key = load_deepseek_api_key()
    if not api_key:
        print("[WARN] DEEPSEEK_API_KEY missing; using local screening fallback")
        return fallback_screen_candidates(candidates, topics)

    topic_text = "\n".join(
        f"{topic['id']}. {topic['name']}" for topic in topics
    )
    candidate_payload = [
        {
            "arxiv_id": paper["arxiv_id"],
            "title": paper["title"],
            "abstract": paper["summary"],
            "candidate_topic_ids": paper.get("candidate_topic_ids", []),
        }
        for paper in candidates
    ]
    prompt = f"""
你是推荐系统论文筛选器。根据论文标题和 arXiv 摘要完成一次低成本筛选。

主题按优先级排列：
{topic_text}

要求：
1. 每篇论文只能分到一个最匹配主题；不属于上述方向时 topic_id=0。
2. 只有论文的主要研究问题、核心方法或核心应用明确属于推荐系统时才能保留。
   如果推荐系统只出现在实验任务列表、背景、相关工作或附带应用中，必须 topic_id=0。
3. relevance_score 为 0-100，衡量论文核心贡献对对应主题的相关性；低于70表示不应进入日报。
4. 代码开源判断只能依据给出的 abstract：
   - 摘要明确写明代码/实现已公开、将公开，或给出 GitHub/GitLab 等仓库链接，code_open_source="是"。
   - 否则 code_open_source="摘要未说明"。不得根据常识或标题猜测。
5. code_url 只能填写摘要中真实出现的代码仓库 URL；没有则为空字符串。
6. 同主题内，代码明确开源的高质量论文应获得更高优先级。
7. 只返回 JSON，不要 Markdown，不要解释。必须为：
{{"papers":[{{"arxiv_id":"...","topic_id":1,"relevance_score":90,
"code_open_source":"是","code_url":""}}]}}

候选论文：
{json.dumps(candidate_payload, ensure_ascii=False)}
""".strip()
    response = requests.post(
        DEEPSEEK_API_URL,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json={
            "model": DEEPSEEK_SCREENING_MODEL,
            "messages": [
                {
                    "role": "system",
                    "content": "Return strict JSON for academic paper screening.",
                },
                {"role": "user", "content": prompt},
            ],
            "response_format": {"type": "json_object"},
            "temperature": 0,
            "max_tokens": 5000,
            "stream": False,
        },
        timeout=180,
    )
    response.raise_for_status()
    payload = response.json()
    content = payload["choices"][0]["message"]["content"]
    usage = payload.get("usage", {})
    print(
        f"[INFO] Screening API tokens | prompt={usage.get('prompt_tokens', '?')} "
        f"completion={usage.get('completion_tokens', '?')} "
        f"total={usage.get('total_tokens', '?')}"
    )
    screening = extract_json_object(content)
    results = {
        str(item.get("arxiv_id", "")).strip(): item
        for item in screening.get("papers", [])
        if str(item.get("arxiv_id", "")).strip()
    }

    topics_by_id = {int(topic["id"]): topic for topic in topics}
    screened = []
    for paper in candidates:
        result = results.get(paper["arxiv_id"])
        if not result:
            continue
        try:
            topic_id = int(result.get("topic_id", 0))
            score = max(0, min(100, int(result.get("relevance_score", 0))))
        except (TypeError, ValueError):
            continue
        if topic_id not in topics_by_id or score < 70:
            continue
        code_status = (
            "是" if result.get("code_open_source") == "是" else "摘要未说明"
        )
        code_url = str(result.get("code_url", "")).strip()
        if code_url and code_url not in paper.get("summary", ""):
            code_url = ""
        screened.append({
            **paper,
            "topic_id": topic_id,
            "topic_name": topics_by_id[topic_id]["name"],
            "screening_score": score,
            "code_open_source": code_status,
            "code_url": code_url,
        })

    print(
        f"[INFO] DeepSeek screened batch of {len(candidates)} candidates; "
        f"kept={len(screened)}"
    )
    return screened


def screen_candidates_with_deepseek(
    candidates: list[dict],
    topics: list[dict],
) -> list[dict]:
    screened = []
    for start in range(0, len(candidates), SCREENING_BATCH_SIZE):
        batch = candidates[start:start + SCREENING_BATCH_SIZE]
        screened.extend(_screen_candidate_batch_with_deepseek(batch, topics))
    return screened


def search_and_select_new_papers(
    topics: list[dict],
    crawled_ids: set[str],
    limit: int,
) -> list[dict]:
    candidates_by_topic: dict[int, list[dict]] = {}
    for index, topic in enumerate(topics):
        topic_id = int(topic["id"])
        try:
            candidates = search_arxiv_papers(topic["query"])
        except Exception as e:
            print(f"[ERROR] Topic {topic_id} search failed: {e}")
            candidates = []
        recent_candidates = [
            paper for paper in candidates
            if paper_is_recent(paper)
        ]
        uncrawled_recent_candidates = [
            paper for paper in recent_candidates
            if paper["arxiv_id"] not in crawled_ids
        ]
        candidates_by_topic[topic_id] = [
            paper for paper in uncrawled_recent_candidates
            if paper_matches_topic(paper, topic)
        ]
        print(
            f"[INFO] Topic {topic_id} candidates | "
            f"retrieved={len(candidates)} recent{RECENT_PAPER_DAYS}={len(recent_candidates)} "
            f"uncrawled={len(uncrawled_recent_candidates)} "
            f"matched={len(candidates_by_topic[topic_id])}"
        )
        if index < len(topics) - 1:
            time.sleep(REQUEST_INTERVAL)

    screening_pool = build_screening_pool(candidates_by_topic, topics)
    print(
        f"[INFO] Sending {len(screening_pool)} title+abstract candidates "
        f"to {DEEPSEEK_SCREENING_MODEL}"
    )
    try:
        screened = screen_candidates_with_deepseek(screening_pool, topics)
    except Exception as e:
        print(f"[WARN] DeepSeek screening failed: {e}; using local fallback")
        screened = fallback_screen_candidates(screening_pool, topics)

    screened_by_topic: dict[int, list[dict]] = {
        int(topic["id"]): [] for topic in topics
    }
    for paper in screened:
        screened_by_topic[int(paper["topic_id"])].append(paper)

    selected = select_topic_papers(screened_by_topic, topics, limit)
    print(f"[INFO] Selected {len(selected)} papers for today's LLM budget ({limit})")
    for paper in selected:
        print(
            f"  - T{paper['topic_id']} [{paper['arxiv_id']}] "
            f"{paper['title'][:70]}"
        )
    return selected


def download_pdf(paper: dict) -> bool:
    pdf_path = PAPERS_DIR / paper["pdf_filename"]
    if pdf_path.exists():
        print(f"[INFO] PDF exists: {paper['pdf_filename']}")
        return True
    try:
        response = requests.get(paper["pdf_url"], timeout=60, stream=True)
        response.raise_for_status()
        with open(pdf_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        print(f"[INFO] Downloaded: {paper['pdf_filename']}")
        return True
    except Exception as e:
        print(f"[ERROR] Download failed {paper['pdf_url']}: {e}")
        return False


def load_or_create_excel() -> openpyxl.Workbook:
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Papers" not in wb.sheetnames:
            wb.create_sheet("Papers")
        ensure_excel_columns(wb["Papers"])
        return wb

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Papers"
    ws.append(EXCEL_COLUMNS)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, _ in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = [15, 45, 28, 35, 12, 18, 70, 60, 34, 14, 35, 20, 12, 25]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    return wb


def ensure_excel_columns(ws: openpyxl.worksheet.worksheet.Worksheet):
    """Add newly introduced persistent columns without rewriting existing data."""
    existing = {
        str(ws.cell(row=1, column=col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=col).value is not None
    }
    for column in EXCEL_COLUMNS:
        if column in existing:
            continue
        new_col = ws.max_column + 1
        cell = ws.cell(row=1, column=new_col, value=column)
        cell.fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        existing[column] = new_col


def migrate_excel_schema():
    if not EXCEL_FILE.exists():
        return
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "Papers" not in wb.sheetnames:
        return
    ws = wb["Papers"]
    before = ws.max_column
    ensure_excel_columns(ws)
    header_index, _ = build_excel_row_index(ws)
    code_col = header_index.get("code_open_source")
    if code_col:
        for row in range(2, ws.max_row + 1):
            if not ws.cell(row=row, column=code_col).value:
                ws.cell(row=row, column=code_col, value="摘要未说明")
    if ws.max_column != before or code_col:
        wb.save(EXCEL_FILE)


def append_to_excel(wb: openpyxl.Workbook, paper: dict):
    ws = wb["Papers"]
    today = date.today().isoformat()
    values = {
        "arxiv_id": paper["arxiv_id"],
        "title": paper["title"],
        "authors": paper["authors"],
        "affiliations": paper.get("affiliations", ""),
        "published_date": paper["published_date"],
        "categories": paper["categories"],
        "abstract": paper["summary"],
        "summary_cn": paper.get("summary_cn", ""),
        "topic_name": paper.get("topic_name", ""),
        "code_open_source": paper.get("code_open_source", "摘要未说明"),
        "code_url": paper.get("code_url", ""),
        "pdf_filename": paper["pdf_filename"],
        "crawled_date": today,
        "notes": "",
    }
    headers = [
        str(ws.cell(row=1, column=col).value)
        for col in range(1, ws.max_column + 1)
    ]
    row = [values.get(header, "") for header in headers]
    ws.append(row)
    last_row = ws.max_row
    for col in range(1, len(row) + 1):
        ws.cell(row=last_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    print(f"[INFO] Appended: {paper['arxiv_id']} - {paper['title'][:40]}...")


def build_excel_row_index(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[dict[str, int], dict[str, int]]:
    """返回 (header_index, arxiv_id -> row_number)。"""
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_index = {str(v): i + 1 for i, v in enumerate(header_row) if v is not None}

    row_index: dict[str, int] = {}
    arxiv_col = header_index.get("arxiv_id")
    if not arxiv_col:
        return header_index, row_index

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=arxiv_col).value
        if val is None:
            continue
        key = str(val).strip()
        if key and key not in row_index:
            row_index[key] = r
    return header_index, row_index


def upsert_to_excel(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_index: dict[str, int],
    row_index: dict[str, int],
    paper: dict,
):
    """按 arxiv_id 更新或插入，避免重复行。"""
    arxiv_id = paper["arxiv_id"]
    today = date.today().isoformat()

    if arxiv_id in row_index:
        target_row = row_index[arxiv_id]
        updates = {
            "title": paper["title"],
            "authors": paper["authors"],
            "published_date": paper["published_date"],
            "categories": paper["categories"],
            "abstract": paper["summary"],
            "topic_name": paper.get("topic_name", ""),
            "code_open_source": paper.get("code_open_source", "摘要未说明"),
            "code_url": paper.get("code_url", ""),
            "pdf_filename": paper["pdf_filename"],
            "crawled_date": today,
        }
        # 仅在新值非空时覆盖，避免把已有结果清空。
        if paper.get("affiliations"):
            updates["affiliations"] = paper.get("affiliations", "")
        if paper.get("summary_cn"):
            updates["summary_cn"] = paper.get("summary_cn", "")

        for key, value in updates.items():
            col = header_index.get(key)
            if col:
                ws.cell(row=target_row, column=col, value=value)

        for col in range(1, ws.max_column + 1):
            ws.cell(row=target_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        print(f"[INFO] Updated: {arxiv_id} - {paper['title'][:40]}...")
        return

    # 不存在则新增
    append_to_excel(wb=ws.parent, paper=paper)
    row_index[arxiv_id] = ws.max_row


def save_excel(wb: openpyxl.Workbook):
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)
    print(f"[INFO] Excel saved: {EXCEL_FILE}")


def export_viewer_json_from_excel():
    """从 papers_record.xlsx 导出 viewer 使用的 papers_data.json。"""
    if not EXCEL_FILE.exists():
        print(f"[WARN] Excel not found, skip viewer export: {EXCEL_FILE}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    if "Papers" not in wb.sheetnames:
        print("[WARN] Sheet 'Papers' not found, skip viewer export")
        return

    ws = wb["Papers"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        print("[WARN] Excel header missing, skip viewer export")
        return

    headers = [str(h) if h is not None else "" for h in header_row]
    index = {name: i for i, name in enumerate(headers)}
    required = EXCEL_COLUMNS
    missing = [c for c in required if c not in index]
    if missing:
        print(f"[WARN] Missing columns in Excel, skip viewer export: {missing}")
        return

    def norm(v: object) -> str:
        if v is None:
            return ""
        return str(v).replace("\n", " ").strip()

    def quality_key(p: dict) -> tuple:
        # 优先保留信息更完整的一行：有中文总结、有单位、文本更长、日期更新。
        return (
            1 if p.get("summary_cn") else 0,
            1 if p.get("affiliations") else 0,
            len(p.get("summary_cn", "")),
            len(p.get("affiliations", "")),
            len(p.get("abstract", "")),
            p.get("crawled_date", ""),
            p.get("published_date", ""),
        )

    papers_by_id: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        paper = {col: norm(row[index[col]]) for col in required}
        if not paper["arxiv_id"]:
            continue
        paper["pdf_url"] = f"https://arxiv.org/pdf/{paper['arxiv_id']}"
        arxiv_id = paper["arxiv_id"]
        old = papers_by_id.get(arxiv_id)
        if old is None or quality_key(paper) > quality_key(old):
            papers_by_id[arxiv_id] = paper

    papers = list(papers_by_id.values())

    papers.sort(
        key=lambda x: (x["crawled_date"], x["published_date"], x["arxiv_id"]),
        reverse=True,
    )

    crawled_dates = sorted({p["crawled_date"] for p in papers if p["crawled_date"]})
    published_dates = sorted({p["published_date"] for p in papers if p["published_date"]})

    payload = {
        "count": len(papers),
        "crawled_date_min": crawled_dates[0] if crawled_dates else "",
        "crawled_date_max": crawled_dates[-1] if crawled_dates else "",
        "published_date_min": published_dates[0] if published_dates else "",
        "published_date_max": published_dates[-1] if published_dates else "",
        "papers": papers,
    }

    VIEWER_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(VIEWER_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"[INFO] Viewer JSON updated: {VIEWER_JSON} (count={len(papers)})")


def load_incomplete_papers_from_excel() -> dict[str, dict]:
    """读取 Excel 中尚未完成 LLM 补全的论文。"""
    if not EXCEL_FILE.exists():
        return {}

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    if "Papers" not in wb.sheetnames:
        return {}

    ws = wb["Papers"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {}

    headers = [str(h) if h is not None else "" for h in header_row]
    index = {name: i for i, name in enumerate(headers)}
    required = EXCEL_COLUMNS
    missing = [c for c in required if c not in index]
    if missing:
        print(f"[WARN] Missing columns in Excel, skip pending check: {missing}")
        return {}

    def norm(v: object) -> str:
        if v is None:
            return ""
        return str(v).replace("\n", " ").strip()

    pending: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        paper = {col: norm(row[index[col]]) for col in required}
        arxiv_id = paper["arxiv_id"]
        if not arxiv_id:
            continue
        if paper["affiliations"] and paper["summary_cn"]:
            continue
        paper["summary"] = paper["abstract"]
        paper["pdf_url"] = f"https://arxiv.org/pdf/{arxiv_id}"
        paper["pdf_local_path"] = str(PAPERS_DIR / paper["pdf_filename"]) if paper["pdf_filename"] else ""
        pending[arxiv_id] = paper
    return pending


def write_llm_output_json(
    papers_to_process: list[dict],
    fresh_downloaded_count: int = 0,
    pending_total_count: int | None = None,
    feishu_msg: str = "",
):
    """输出当前待处理状态，供 Hermes agent 继续执行或安全重试。"""
    if pending_total_count is None:
        pending_total_count = len(papers_to_process)
    output = {
        "date": date.today().isoformat(),
        "new_count": fresh_downloaded_count,
        "pending_count": len(papers_to_process),
        "pending_total_count": pending_total_count,
        "daily_llm_limit": DAILY_LLM_LIMIT,
        "excel_file": str(EXCEL_FILE),
        "papers_dir": str(PAPERS_DIR),
        "new_papers": papers_to_process,
        "papers_to_process": papers_to_process,
        "feishu_msg": feishu_msg,
    }
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)


def sync_pending_state_from_excel(refresh_output_json: bool = True) -> list[dict]:
    """
    根据 Excel 当前状态重建 pending_llm_ids.txt。
    可选同时刷新 new_papers.json，避免 agent 重试时继续读取旧待处理列表。
    """
    incomplete_excel_papers = load_incomplete_papers_from_excel()
    all_pending_papers = [
        incomplete_excel_papers[arxiv_id]
        for arxiv_id in sorted(incomplete_excel_papers)
    ]
    papers_to_process = all_pending_papers[:DAILY_LLM_LIMIT]
    save_pending_llm_ids({p["arxiv_id"] for p in all_pending_papers})
    if refresh_output_json:
        write_llm_output_json(
            papers_to_process=papers_to_process,
            pending_total_count=len(all_pending_papers),
        )
    return papers_to_process


# ==================== 主流程 ====================

def main():
    migrate_excel_schema()

    if len(sys.argv) > 1 and sys.argv[1] == "--sync-pending-state":
        papers_to_process = sync_pending_state_from_excel(refresh_output_json=True)
        pending_total = len(load_incomplete_papers_from_excel())
        print(
            f"[INFO] Pending LLM state synced from Excel | "
            f"next_batch={len(papers_to_process)} remaining_total={pending_total} "
            f"output_json={OUTPUT_JSON}"
        )
        return

    print("=" * 60)
    print(f"[START] arxiv Monitor | {datetime.now().isoformat()}")
    print("=" * 60)

    PAPERS_DIR.mkdir(parents=True, exist_ok=True)

    # 加载已爬取 ID（查重）：Excel 为主，txt 为兜底缓存。
    crawled_ids_txt = load_crawled_ids()
    crawled_ids_excel = load_excel_ids()
    crawled_ids = crawled_ids_txt | crawled_ids_excel
    incomplete_excel_papers = load_incomplete_papers_from_excel()
    pending_ids_file = load_pending_llm_ids()
    all_pending_ids = set(incomplete_excel_papers)
    save_pending_llm_ids(all_pending_ids)
    print(
        f"[INFO] crawled IDs loaded | txt={len(crawled_ids_txt)} "
        f"excel={len(crawled_ids_excel)} merged={len(crawled_ids)}"
    )
    print(
        f"[INFO] pending LLM IDs loaded | file={len(pending_ids_file)} "
        f"excel_incomplete={len(incomplete_excel_papers)}"
    )

    # 失败重试优先占用每日 LLM 预算，剩余名额才用于抓取新论文。
    pending_batch_ids = sorted(all_pending_ids)[:DAILY_LLM_LIMIT]
    remaining_new_slots = DAILY_LLM_LIMIT - len(pending_batch_ids)
    print(
        f"[INFO] Daily LLM budget | limit={DAILY_LLM_LIMIT} "
        f"retry_slots={len(pending_batch_ids)} new_slots={remaining_new_slots}"
    )

    topics = load_search_topics()
    new_papers = search_and_select_new_papers(
        topics=topics,
        crawled_ids=crawled_ids,
        limit=remaining_new_slots,
    ) if remaining_new_slots else []

    # 下载 PDF + 更新 ID
    downloaded = []
    for index, paper in enumerate(new_papers):
        ok = download_pdf(paper)
        if ok:
            downloaded.append({**paper, "pdf_downloaded": True})
        if index < len(new_papers) - 1:
            time.sleep(REQUEST_INTERVAL)

    if downloaded:
        # 保存 Excel（summary_cn 和 affiliations 暂留空，等 LLM 填入）
        wb = load_or_create_excel()
        ws = wb["Papers"]
        header_index, row_index = build_excel_row_index(ws)
        for paper in downloaded:
            upsert_to_excel(ws, header_index, row_index, paper)
        save_excel(wb)

        # 批量写入 crawled_ids
        save_crawled_ids_batch([p["arxiv_id"] for p in downloaded])

    incomplete_excel_papers = load_incomplete_papers_from_excel()
    all_pending_ids = set(incomplete_excel_papers)
    save_pending_llm_ids(all_pending_ids)

    # Preserve retry-first ordering, then append today's selected papers.
    batch_ids = [
        arxiv_id for arxiv_id in pending_batch_ids
        if arxiv_id in incomplete_excel_papers
    ]
    batch_ids.extend(
        paper["arxiv_id"] for paper in downloaded
        if paper["arxiv_id"] in incomplete_excel_papers
    )
    batch_ids = list(dict.fromkeys(batch_ids))[:DAILY_LLM_LIMIT]
    selected_metadata = {paper["arxiv_id"]: paper for paper in downloaded}
    papers_to_process = []
    for arxiv_id in batch_ids:
        paper = incomplete_excel_papers[arxiv_id]
        metadata = selected_metadata.get(arxiv_id, {})
        papers_to_process.append({
            **paper,
            **{
                key: metadata[key]
                for key in ("topic_id", "topic_name", "selection_score")
                if key in metadata
            },
        })

    if not papers_to_process:
        # 无新论文，且无待补全论文
        output = {
            "date": date.today().isoformat(),
            "new_count": 0,
            "pending_count": 0,
            "pending_total_count": 0,
            "daily_llm_limit": DAILY_LLM_LIMIT,
            "new_papers": [],
            "papers_to_process": [],
            "feishu_msg": f"✅ 今日（{date.today().isoformat()}）未发现新的生成式推荐论文。",
        }
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        export_viewer_json_from_excel()
        print("[INFO] No new papers and no pending LLM tasks. Output JSON written.")
        return

    # 输出 JSON（供 hermes agent 读取并做 LLM summarization）
    write_llm_output_json(
        papers_to_process=papers_to_process,
        fresh_downloaded_count=len(downloaded),
        pending_total_count=len(all_pending_ids),
    )

    print(f"[INFO] Output JSON: {OUTPUT_JSON}")
    print(
        f"[INFO] fresh_downloaded={len(downloaded)} "
        f"llm_batch={len(papers_to_process)} "
        f"pending_total={len(all_pending_ids)}. Awaiting LLM summarization..."
    )

    print("\n" + "=" * 60)
    print("[LLM_SUMMARIZATION_REQUIRED]")
    print("=" * 60)
    print(f"JSON file: {OUTPUT_JSON}")
    print(f"Fresh downloads: {len(downloaded)}")
    print(f"Papers awaiting LLM completion: {len(papers_to_process)}")
    for p in papers_to_process:
        print(f"  - [{p['arxiv_id']}] {p['title'][:50]}... | PDF: {p['pdf_filename']}")
    print("=" * 60)
    print("请用 LLM（claude/gpt）完成以下步骤：")
    print("  1. 读取每个 PDF 的前两页，提取作者单位（affiliations）")
    print("  2. 对每篇论文的 abstract 生成 150 字以内的中文总结（summary_cn）")
    print("  3. 将结果更新回 papers_record.xlsx 对应行")
    print("  4. 重建 viewer/papers_data.json")
    print("  5. 构建飞书 Markdown 消息并输出（cronjob 自动投递到飞书）")
    print("=" * 60)


if __name__ == "__main__":
    main()
