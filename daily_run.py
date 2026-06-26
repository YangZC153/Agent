#!/usr/bin/env python3
"""Deterministic daily pipeline for no-agent Hermes cron execution."""

from __future__ import annotations

import contextlib
import io
import json
import os
import subprocess
import time
from datetime import datetime
from pathlib import Path

import fitz
import openpyxl
import requests

from investment_advice import dca as investment_dca
import monitor


PROJECT_DIR = Path(__file__).resolve().parent
MODEL = "deepseek-v4-flash"
MAX_PDF_TEXT_CHARS = 12000
API_ATTEMPTS = 2
DEEPSEEK_TIMEOUT = 120
RUN_LOG = PROJECT_DIR / "logs" / f"daily_run_{monitor.date.today().isoformat()}.log"
DEPLOY_MODE_FILE = PROJECT_DIR / ".deploy_mode"
SILENT_MARKER = "[SILENT]"


def log_event(message: str) -> None:
    RUN_LOG.parent.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().isoformat(timespec="seconds")
    with RUN_LOG.open("a", encoding="utf-8") as handle:
        handle.write(f"[{timestamp}] {message}\n")


class LogCapture:
    def __init__(self, buffer: io.StringIO):
        self.buffer = buffer
        RUN_LOG.parent.mkdir(parents=True, exist_ok=True)

    def write(self, text: str) -> int:
        self.buffer.write(text)
        with RUN_LOG.open("a", encoding="utf-8") as handle:
            handle.write(text)
        return len(text)

    def flush(self) -> None:
        return None


def extract_first_pages(pdf_path: str, pages: int = 2) -> str:
    document = fitz.open(pdf_path)
    try:
        text = "\n\n".join(
            document[index].get_text("text")
            for index in range(min(pages, document.page_count))
        )
    finally:
        document.close()
    return text[:MAX_PDF_TEXT_CHARS]


def call_deepseek_json(prompt: str) -> dict:
    api_key = monitor.load_deepseek_api_key()
    if not api_key:
        raise RuntimeError("DEEPSEEK_API_KEY is not configured")

    last_error = None
    for attempt in range(1, API_ATTEMPTS + 1):
        try:
            response = requests.post(
                monitor.DEEPSEEK_API_URL,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": MODEL,
                    "messages": [
                        {
                            "role": "system",
                            "content": "Return strict JSON only.",
                        },
                        {"role": "user", "content": prompt},
                    ],
                    "response_format": {"type": "json_object"},
                    "temperature": 0,
                    "max_tokens": 1800,
                    "stream": False,
                },
                timeout=DEEPSEEK_TIMEOUT,
            )
            response.raise_for_status()
            content = response.json()["choices"][0]["message"]["content"]
            result = monitor.extract_json_object(content)
            if not result:
                raise ValueError("empty JSON response")
            return result
        except Exception as exc:
            last_error = exc
            if attempt < API_ATTEMPTS:
                time.sleep(2 ** attempt)
    raise RuntimeError(f"DeepSeek request failed after {API_ATTEMPTS} attempts: {last_error}")


def process_paper(paper: dict) -> dict:
    pdf_text = extract_first_pages(paper["pdf_local_path"])
    prompt = f"""
根据给定 arXiv 论文信息返回 JSON：
{{
  "affiliations": "作者单位，多个单位用英文分号分隔",
  "summary_cn": "90-150个中文字符的单段总结"
}}

要求：
1. affiliations 仅从 PDF 前两页提取。删除邮箱、URL、脚注编号、正文句子和公式；无法确定时写“未找到单位信息”。
2. summary_cn 仅根据 abstract，总结方法核心、主要贡献和关键结果。不要分点，不要模板化。
3. 只返回 JSON。

标题：{paper["title"]}
Abstract：{paper["summary"]}
PDF前两页文本：
{pdf_text}
""".strip()
    result = call_deepseek_json(prompt)
    affiliations = str(result.get("affiliations", "")).replace("\n", " ").strip()
    summary_cn = str(result.get("summary_cn", "")).replace("\n", " ").strip()
    if not affiliations:
        affiliations = "未找到单位信息"
    if not summary_cn:
        raise ValueError(f"empty summary_cn for {paper['arxiv_id']}")
    return {
        "affiliations": affiliations,
        "summary_cn": summary_cn,
    }


def normalize_arxiv_id(value: object) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    if "." in text:
        prefix, suffix = text.split(".", 1)
        text = f"{prefix}.{suffix.rstrip('0')}"
    return text


def update_excel(results: dict[str, dict]) -> None:
    workbook = openpyxl.load_workbook(monitor.EXCEL_FILE)
    sheet = workbook["Papers"]
    headers = {
        str(sheet.cell(row=1, column=col).value): col
        for col in range(1, sheet.max_column + 1)
    }
    normalized_results = {
        normalize_arxiv_id(arxiv_id): result
        for arxiv_id, result in results.items()
    }
    for row in range(2, sheet.max_row + 1):
        arxiv_id = normalize_arxiv_id(
            sheet.cell(row=row, column=headers["arxiv_id"]).value
        )
        result = normalized_results.get(arxiv_id)
        if not result:
            continue
        sheet.cell(
            row=row,
            column=headers["affiliations"],
            value=result["affiliations"],
        )
        sheet.cell(
            row=row,
            column=headers["summary_cn"],
            value=result["summary_cn"],
        )
    workbook.save(monitor.EXCEL_FILE)


def format_report(
    papers: list[dict],
    results: dict[str, dict],
    failures: list[str] | None = None,
    viewer_published: bool = False,
    search_failures: list[dict] | None = None,
) -> str:
    failures = failures or []
    search_failures = search_failures or []
    if not papers:
        if search_failures:
            lines = [
                f"论文日报 | {monitor.date.today().isoformat()}",
                "今日 arXiv 检索未完全成功，部分主题因超时或限流没有返回结果。",
                "为避免误判为“无新论文”，本次不发布空论文日报；系统会在下次定时运行继续检索。",
                "",
                "未完成主题：",
            ]
            lines.extend(format_search_failures(search_failures))
            return "\n".join(lines).strip()
        return f"今日（{monitor.date.today().isoformat()}）未发现新的相关论文。"

    successful_papers = [
        paper for paper in papers if paper["arxiv_id"] in results
    ]
    lines = [
        f"论文日报 | {monitor.date.today().isoformat()}",
        f"共发现 {len(papers)} 篇新论文，已完成总结 {len(successful_papers)} 篇",
        "",
    ]
    for index, paper in enumerate(successful_papers, 1):
        result = results[paper["arxiv_id"]]
        lines.extend([
            f"{index}. {paper['title']}",
            f"主题: {paper.get('topic_name') or '未分类'}",
            f"arXiv: {paper['arxiv_id']} | {paper['published_date']}",
            f"作者: {paper['authors']}",
            f"单位: {result['affiliations']}",
            f"代码开源: {paper.get('code_open_source') or '摘要未说明'}",
        ])
        if paper.get("code_url"):
            lines.append(f"代码: {paper['code_url']}")
        lines.extend([
            f"PDF: {paper['pdf_url']}",
            f"中文总结: {result['summary_cn']}",
            "",
        ])
    if search_failures:
        lines.extend([
            "检索提示：部分主题检索失败，本次结果可能不完整。",
            *format_search_failures(search_failures),
            "",
        ])
    if failures:
        lines.extend([
            "以下论文处理失败，将在下次自动重试：",
            *failures,
            "",
        ])
    if viewer_published:
        lines.append("完整列表: https://yangzc153.github.io/Agent/")
    elif failures:
        lines.append("网页未发布本次结果，避免展示未补全的论文记录。")
    else:
        lines.append("结果已直接推送；当前为本地模式，未发布到 GitHub Pages 展示页。")
    return "\n".join(lines).strip()


def format_search_failures(search_failures: list[dict]) -> list[str]:
    lines = []
    for failure in search_failures:
        topic_id = failure.get("topic_id", "")
        topic_name = failure.get("topic_name", "") or "未命名主题"
        error = str(failure.get("error", "")).strip()
        if len(error) > 140:
            error = f"{error[:137]}..."
        lines.append(f"- T{topic_id} {topic_name}: {error}")
    return lines


def format_investment_report() -> str:
    try:
        result = investment_dca.generate_investment_advice()
        return investment_dca.format_public_report(result)
    except Exception as exc:
        log_event(f"investment advice failed: {type(exc).__name__}: {exc}")
        investment_dca.log_investment_event(
            f"ERROR daily_run investment advice failed: {type(exc).__name__}: {exc}"
        )
        return investment_dca.format_public_report(
            investment_dca.safe_fallback_advice(exc)
        )


def direct_send_enabled() -> bool:
    value = os.getenv("DAILY_RUN_DIRECT_SEND", "").strip().lower()
    return value in {"1", "true", "yes", "on"}


def hermes_bin() -> Path:
    configured = os.getenv("HERMES_BIN", "").strip()
    if configured:
        return Path(configured)
    return Path(monitor.sys.executable).with_name("hermes")


def send_direct_message(message: str) -> None:
    target = os.getenv("DAILY_RUN_SEND_TARGET", "feishu").strip() or "feishu"
    command = [
        str(hermes_bin()),
        "send",
        "--to",
        target,
        "--file",
        "-",
        "--quiet",
    ]
    result = subprocess.run(
        command,
        cwd=PROJECT_DIR,
        input=message,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        timeout=180,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"hermes send failed for {target}: {result.stdout.strip()}"
        )
    log_event(f"direct message delivered to {target}")


def deliver_daily_messages(arxiv_report: str) -> None:
    investment_report = format_investment_report()
    if direct_send_enabled():
        sent_arxiv = False
        sent_investment = False
        try:
            send_direct_message(arxiv_report)
            sent_arxiv = True
            send_direct_message(investment_report)
            sent_investment = True
        except Exception as exc:
            log_event(f"direct delivery failed: {type(exc).__name__}: {exc}")
            fallback_parts = []
            if not sent_arxiv:
                fallback_parts.append(arxiv_report)
            if not sent_investment:
                fallback_parts.append(investment_report)
            if fallback_parts:
                print("\n\n".join(fallback_parts))
            else:
                print(SILENT_MARKER)
            return
        print(SILENT_MARKER)
        return

    print(arxiv_report)
    print()
    print(investment_report)


def run_monitor() -> dict:
    buffer = io.StringIO()
    log_event("monitor.py output begins")
    with contextlib.redirect_stdout(LogCapture(buffer)):
        monitor.main()
    log_event("monitor.py output ends")
    return json.loads(monitor.OUTPUT_JSON.read_text(encoding="utf-8"))


def run_command(command: list[str]) -> None:
    result = subprocess.run(
        command,
        cwd=PROJECT_DIR,
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        timeout=900,
    )
    log_event(f"command {' '.join(command)} output:\n{result.stdout.strip()}")


def get_deploy_mode() -> str:
    if not DEPLOY_MODE_FILE.exists():
        return "local"
    mode = DEPLOY_MODE_FILE.read_text(encoding="utf-8").strip().lower()
    return mode if mode in {"local", "pages"} else "local"


def should_publish_viewer() -> bool:
    override = os.getenv("PUBLISH_VIEWER", "").strip().lower()
    if override:
        return override in {"1", "true", "yes", "on"}
    return get_deploy_mode() == "pages"


def main() -> int:
    log_event("daily_run started")
    try:
        payload = run_monitor()
        papers = payload.get("papers_to_process", [])[:monitor.DAILY_LLM_LIMIT]
        search_failures = payload.get("search_failures", [])
        log_event(f"papers_to_process={len(papers)}")
        if search_failures:
            log_event(f"search_failures={json.dumps(search_failures, ensure_ascii=False)}")
        if not papers:
            deliver_daily_messages(format_report([], {}, search_failures=search_failures))
            log_event("daily_run finished: no papers")
            return 0

        results = {}
        failures = []
        for paper in papers:
            try:
                log_event(f"processing {paper['arxiv_id']}")
                results[paper["arxiv_id"]] = process_paper(paper)
                log_event(f"processed {paper['arxiv_id']}")
            except Exception as exc:
                failures.append(f"{paper['arxiv_id']}: {exc}")
                log_event(f"failed {paper['arxiv_id']}: {exc}")

        if results:
            update_excel(results)

        with contextlib.redirect_stdout(io.StringIO()):
            monitor.sync_pending_state_from_excel(refresh_output_json=True)

        if failures:
            deliver_daily_messages(format_report(
                papers,
                results,
                failures,
                viewer_published=False,
                search_failures=search_failures,
            ))
            log_event("daily_run finished with partial failures")
            return 0

        run_command([
            str(Path(monitor.sys.executable)),
            "viewer/build_data.py",
        ])
        viewer_published = False
        if should_publish_viewer():
            run_command(["bash", "scripts/publish_viewer.sh"])
            viewer_published = True
        else:
            log_event(
                "viewer publish skipped: "
                f"deploy_mode={get_deploy_mode()} "
                f"PUBLISH_VIEWER={os.getenv('PUBLISH_VIEWER', 'unset')}"
            )
        deliver_daily_messages(format_report(
            papers,
            results,
            viewer_published=viewer_published,
            search_failures=search_failures,
        ))
        log_event("daily_run finished successfully")
        return 0
    except Exception as exc:
        log_event(f"daily_run failed: {type(exc).__name__}: {exc}")
        deliver_daily_messages(
            f"论文日报运行失败，将在下次自动重试。\n{type(exc).__name__}: {exc}"
        )
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
